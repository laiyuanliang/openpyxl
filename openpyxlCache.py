#插入一张图片

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.drawing.image import Image

wb = load_workbook("sample.xlsx")

img = Image('girl.jpg')
sheet = wb['mysheet']
sheet.add_image(img,'B3')

wb.save('sample.xlsx')

#插入一个摘要
from openpyxl import Workbook
from openpyxl.comments import Comment

wb = Workbook()
ws = wb.active
ws.title = 'commenttest'

ws['F5'].comment = Comment('this is test for comment','author')

wb.save('commenttest.xlsx')

#在工作簿的每一个工作表内插入一个摘要
from openpyxl import load_workbook
from openpyxl.comments import Comment

wb = load_workbook('ymz.xlsx')
sheet_numbers = len(wb.sheetnames)
print(sheet_numbers)
for i in range(sheet_numbers):
    ws = wb.worksheets[i]
    ws['B106'].comment = Comment("I'm tring to write comment to multiple sheet", "author")

wb.save('ymz.xlsx')


from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter

wb = Workbook()
ws1 = wb.active
ws1.title = "firstSheet"

for row in range(1,40):
    ws1.append(range(600))

ws2 = wb.create_sheet(title = "pi")
ws2["F5"] = 3.14

ws3 = wb.create_sheet("Data")
for row in range(5,15):
    for col in range(6,10):
        ws3.cell(column=col, row=row, value = "{0}".format(get_column_letter(col)))

wb.save("simple.xlsx")

#闭包
def adder(x):
    def wrapper(y):
        return x+y
    return wrapper()

adder_5 = adder(5)
print(adder_5(2))
