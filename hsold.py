# -*- coding: utf-8 -*-
import io,os
from openpyxl import Workbook
from openpyxl.compat import range
from openpyxl import load_workbook
import datetime,time
import traceback

sourcefile = u'/root/ms.xlsx' 
outputfile = u'/root/msout.xlsx'  

index = 1;

wb = load_workbook(sourcefile)
sheet = wb.active

wbout = Workbook()
ws = wbout.active
ws.cell(column=1, row=1).value = u"序号"
ws.cell(column=2, row=1).value = u"姓名"
ws.cell(column=3, row=1).value = u"手机号"
ws.cell(column=4, row=1).value = u"身份证号"
ws.cell(column=5, row=1).value = u"创建时间"
ws.cell(column=6, row=1).value = u"信贷平台注册详情(如有多条命中，以分号分隔)"
ws.cell(column=7, row=1).value = u"注册次数"
ws.cell(column=8, row=1).value = u"注册次数（银行）"
ws.cell(column=9, row=1).value = u"注册次数（非银行）"
ws.cell(column=10, row=1).value = u"贷款申请详情"
ws.cell(column=11, row=1).value = u"申请次数"
ws.cell(column=12, row=1).value = u"申请次数（银行）"
ws.cell(column=13, row=1).value = u"申请次数（非银行）"
ws.cell(column=14, row=1).value = u"申请金额区间（0-0.2w）"
ws.cell(column=15, row=1).value = u"申请金额区间(0.2-0.5w)"
ws.cell(column=16, row=1).value = u"申请金额区间(0.5-1w)"
ws.cell(column=17, row=1).value = u"申请金额区间(1-3w)"
ws.cell(column=18, row=1).value = u"申请金额区间(3-5w)"
ws.cell(column=19, row=1).value = u"申请金额区间(5-10w)"
ws.cell(column=20, row=1).value = u"申请金额区间(10w以上)"
ws.cell(column=21, row=1).value = u"最大申请金额区间"
ws.cell(column=22, row=1).value = u"贷款放款详情"
ws.cell(column=23, row=1).value = u"放款次数"
ws.cell(column=24, row=1).value = u"放款次数(银行)"
ws.cell(column=25, row=1).value = u"放款次数(非银行)"
ws.cell(column=26, row=1).value = u"放款金额区间(0-0.2w)"
ws.cell(column=27, row=1).value = u"放款金额区间(0.2-0.5w)"
ws.cell(column=28, row=1).value = u"放款金额区间(0.5-1w)"
ws.cell(column=29, row=1).value = u"放款金额区间(1-3w)"
ws.cell(column=30, row=1).value = u"放款金额区间(3-5w)"
ws.cell(column=31, row=1).value = u"放款金额区间(5-10w)"
ws.cell(column=32, row=1).value = u"放款金额区间(10w以上)"
ws.cell(column=33, row=1).value = u"最大放款金额区间"
ws.cell(column=34, row=1).value = u"贷款驳回详情"
ws.cell(column=35, row=1).value = u"驳回次数"
ws.cell(column=36, row=1).value = u"驳回次数(银行)"
ws.cell(column=37, row=1).value = u"驳回次数(非银行)"

writeIndex = 2
try:
    for row in sheet.rows:
        if index > 1:
            serialNO = row[0].value
            if serialNO <> None:
                name = row[1].value
                phoneNO = row[2].value
                IDNO = row[3].value
                ApplyTime = row[4].value
                reg_info = row[5].value
                app_info = row[6].value
                loan_info = row[7].value
                rej_info = row[8].value

                #注册统计
                if reg_info != None:
                    reg = len(re.findall(u'平台类型',reg_info))
                    reg_Bank = len(re.findall(u'平台类型:银行',reg_info))
                    reg_NoBank = len(re.findall(u'平台类型:非银行',reg_info))
                else:
                    reg, reg_Bank, reg_NoBank = 0 

                #申请统计
                if app_info != None:
                    app = len(re.findall(u'平台类型', app_info))
                    app_Bank = len(re.findall(u'平台类型:银行', app_info))
                    app_NoBank = len(re.findall(u'平台类型:非银行', app_info))
                    appMount_0t2k = len(re.findall(u'0W～0.2W', app_info))
                    appMount_2t5k = len(re.findall(u'0.2W～0.5W', app_info))
                    appMount_5t10k = len(re.findall(u'0.5W～1W', app_info))
                    appMount_10t30k = len(re.findall(u'1W～3W', app_info))
                    appMount_30t50k = len(re.findall(u'3W～5W', app_info))
                    appMount_50t100k = len(re.findall(u'5W～10W', app_info))
                    appMount_100kt = len(re.findall(u'10W以上', app_info))

                    app_Mount = re.findall('\d*\.*\d*[W][～]\d*\.*\d*[W]|10W以上', app_info)
                    max_app_zone = max_mount(app_Mount)
                else:
                    app, app_Bank, app_NoBank, appMount_0t2k, appMount_2t5k, appMount_5t10k, appMount_10t30k, appMount_30t50k,appMount_50t100k, appMount_100kt, max_app_zone = 0

                #放款统计
                if loan_info != None:
                    loan = len(re.findall(u'平台类型', loan_info))
                    loan_Bank = len(re.findall(u'平台类型:银行', loan_info))
                    loan_NoBank = len(re.findall(u'平台类型:非银行', loan_info))
                    loanMount_0t2k = len(re.findall(u'0W～0.2W', loan_info))
                    loanMount_2t5k = len(re.findall(u'0.2W～0.5W', loan_info))
                    loanMount_5t10k = len(re.findall(u'0.5W～1W', loan_info))
                    loanMount_10t30k = len(re.findall(u'1W～3W', loan_info))
                    loanMount_30t50k = len(re.findall(u'3W～5W', loan_info))
                    loanMount_50t100k = len(re.findall(u'5W～10W', loan_info))
                    loanMount_100kt = len(re.findall(u'10W以上', loan_info))

                    loan_Mount = re.findall('\d*\.*\d*[W][～]\d*\.*\d*[W]|10W以上', loan_info)
                    max_loan_zone = max_mount(loan_Mount)
                else:
                    loan, loan_Bank, loan_NoBank, loanMount_0t2k, loanMount_2t5k, loanMount_5t10k, loanMount_10t30k, loanMount_30t50k, loanMount_50t100k, loanMount_100kt, max_loan_zone = 0

                #驳回统计
                if reg_info != None:
                    rej = len(re.findall(u'平台类型',rej_info))
                    rej_Bank = len(re.findall(u'平台类型:银行',rej_info))
                    rej_NoBank = len(re.findall(u'平台类型:非银行',rej_info))
                else:
                    rej, rej_Bank, reg_NoBank = 0



                #Write File
                    ws.cell(column=1, row=writeIndex).value = serialNO
                    ws.cell(column=2, row=writeIndex).value = name
                    ws.cell(column=3, row=writeIndex).value = phoneNO
                    ws.cell(column=4, row=writeIndex).value = IDNO
                    ws.cell(column=5, row=writeIndex).value = ApplyTime
                    ws.cell(column=6, row=writeIndex).value = platformStr

                    ApplicationStr = ''
                    for Application in outApplicationinfo:
                        ApplicationStr = Application + '\r' + ApplicationStr
                    ws.cell(column=7, row=writeIndex).value = ApplicationStr

                    LoansStr = ''
                    for Loans in outLoansinfo:
                        LoansStr = Loans + '\r' + LoansStr
                    ws.cell(column=8, row=writeIndex).value = LoansStr

                    RejectStr = ''
                    for Reject in outRejectinfo:
                        RejectStr = Reject + '\r' + RejectStr
                    ws.cell(column=9, row=writeIndex).value = RejectStr

                    writeIndex = writeIndex + 1
            # break
        index += 1
        if index%1000 == 0:
            print index


        #写文件

except Exception , e:
    print traceback.print_exc()
    print u"在第%d行出错"%(index)
finally:
    wbout.save(outputfile)
#coding: utf-8
import re

def max_mount(zone_list):
    num_zone = []
    zone_dic = {'0':'0W～0.2W', '0.2':'0.2W～0.5W', '0.5':'0.5W～1W', '1':'1W～3W', '3':'3W～5W', '5':'5W～10W', '10':'10W以上'}
    for zone in zone_list:
        num = zone.split('W')[0]
        num_zone.append(num)
    ordered_num_zone = sorted(num_zone)
    max_zone = zone_dic[ordered_num_zone[-1]]
    return max_zone

reg_info = u'平台类型:银行,平台代码:GEO_0000002633,注册时间:2016-08-12 06:59:57平台类型:非银行,平台代码:GEO_0000184968,注册时间:2016-08-12 06:59:57平台类型:银行,平台代码:GEO_0000001806,注册时间:2016-08-12 06:59:57平台类型:银行,平台代码:GEO_0000000685,注册时间:2016-08-12 06:59:57'
app_info = u'平台类型:银行,平台代码:GEO_0000001806,申请时间:2016-08-12 06:59:57,申请金额区间:0W～0.2W平台类型:银行,平台代码:GEO_0000002633,申请时间:2016-08-12 06:59:57,申请金额区间:0.5W～1W平台类型:银行,平台代码:GEO_0000000685,申请时间:2016-08-12 06:59:57,申请金额区间:10W以上平台类型:非银行,平台代码:GEO_0000184968,申请时间:2016-08-12 06:59:57,申请金额区间:3W～5W'
loan_info = u'平台类型:银行,平台代码:GEO_0000001806,放款时间:2016-08-12 06:59:57,放款金额区间:0.2W～0.5W平台类型:非银行,平台代码:GEO_0000184968,放款时间:2016-08-12 06:59:57,放款金额区间:0.2W～0.5W平台类型:银行,平台代码:GEO_0000000685,放款时间:2016-08-12 06:59:57,放款金额区间:0.2W～0.5W平台类型:银行,平台代码:GEO_0000002633,放款时间:2016-08-12 06:59:57,放款金额区间:0.2W～0.5W'
rej_info = u'平台类型:非银行,平台代码:GEO_0000001668,驳回时间:2017-01-30 12:10:01平台类型:非银行,平台代码:GEO_0000001668,驳回时间:2017-03-08 11:20:00'

#注册统计
reg = len(re.findall(u'平台类型',reg_info))
reg_Bank = len(re.findall(u'平台类型:银行',reg_info))
reg_NoBank = len(re.findall(u'平台类型:非银行',reg_info))

#申请统计
app = len(re.findall(u'平台类型', app_info))
app_Bank = len(re.findall(u'平台类型:银行', app_info))
app_NoBank = len(re.findall(u'平台类型:非银行', app_info))
appMount_0t2k = len(re.findall(u'0W～0.2W', app_info))
appMount_2t5k = len(re.findall(u'0.2W～0.5W', app_info))
appMount_5t10k = len(re.findall(u'0.5W～1W', app_info))
appMount_10t30k = len(re.findall(u'1W～3W', app_info))
appMount_30t50k = len(re.findall(u'3W～5W', app_info))
appMount_50t100k = len(re.findall(u'5W～10W', app_info))
appMount_100kt = len(re.findall(u'10W以上', app_info))

app_Mount = re.findall('\d*\.*\d*[W][～]\d*\.*\d*[W]|10W以上', app_info)
max_app_zone = max_mount(app_Mount)

#放款统计
loan = len(re.findall(u'平台类型', loan_info))
loan_Bank = len(re.findall(u'平台类型:银行', loan_info))
loan_NoBank = len(re.findall(u'平台类型:非银行', loan_info))
loanMount_0t2k = len(re.findall(u'0W～0.2W', loan_info))
loanMount_2t5k = len(re.findall(u'0.2W～0.5W', loan_info))
loanMount_5t10k = len(re.findall(u'0.5W～1W', loan_info))
loanMount_10t30k = len(re.findall(u'1W～3W', loan_info))
loanMount_30t50k = len(re.findall(u'3W～5W', loan_info))
loanMount_50t100k = len(re.findall(u'5W～10W', loan_info))
loanMount_100kt = len(re.findall(u'10W以上', loan_info))

loan_Mount = re.findall('\d*\.*\d*[W][～]\d*\.*\d*[W]|10W以上', loan_info)
max_loan_zone = max_mount(loan_Mount)

#驳回统计
rej = len(re.findall(u'平台类型',rej_info))
rej_Bank = len(re.findall(u'平台类型:银行',rej_info))
rej_NoBank = len(re.findall(u'平台类型:非银行',rej_info))
